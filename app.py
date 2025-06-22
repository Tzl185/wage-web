import streamlit as st
import os
import pandas as pd
from openpyxl import load_workbook
import zipfile
import tempfile
import xlrd

st.set_page_config(page_title="工资表处理工具", layout="centered")

def convert_xls_to_xlsx(xls_path):
    df = pd.read_excel(xls_path, header=None, engine='xlrd')
    new_path = xls_path + "x"
    df.to_excel(new_path, header=False, index=False, engine='openpyxl')
    return new_path

def get_merged_headers(path, header_row=4):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    merged = {}
    for rng in ws.merged_cells.ranges:
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                merged[(row, col)] = ws.cell(rng.min_row, rng.min_col).value
    headers = []
    for col in range(1, ws.max_column + 1):
        top = merged.get((header_row - 1, col)) or ws.cell(header_row - 1, col).value
        bottom = merged.get((header_row, col)) or ws.cell(header_row, col).value
        parts = [str(p).strip() for p in [top, bottom] if p and p != top]
        headers.append("-".join(parts) if parts else f"列{col}")
    return headers

def process_file_a(folder_path, output_file="文件A_汇总结果.xlsx", log_area=None):
    all_data = []
    all_values = {}
    for root, _, files in os.walk(folder_path):
        for fname in files:
            if fname.endswith(('.xls', '.xlsx')) and not fname.startswith('~$'):
                fpath = os.path.join(root, fname)
                try:
                    if log_area: log_area.write(f"\n📄 正在处理: {fname}\n")
                    if fpath.endswith(".xls"):
                        fpath = convert_xls_to_xlsx(fpath)
                        if log_area: log_area.write("📎 已转换为 .xlsx\n")
                    headers = get_merged_headers(fpath)
                    df = pd.read_excel(fpath, header=3, engine='openpyxl')
                    df.columns = headers
                    budget_col = headers[1]
                    wage_cols = headers[16:30]
                    df_filtered = df[[budget_col] + wage_cols]
                    df_filtered[wage_cols] = df_filtered[wage_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
                    df_grouped = df_filtered.groupby(budget_col).sum()
                    for unit, row in df_grouped.iterrows():
                        for wage_type in wage_cols:
                            value = row[wage_type]
                            name = str(wage_type).strip()
                            name = name.replace("绩效工资", "基础性绩效")
                            name = name.replace("行政医疗", "职工基本医疗（行政）")
                            name = name.replace("事业医疗", "基本医疗（事业）")
                            name = name.replace("医疗保险", "基本医疗")
                            all_values[(str(unit).strip(), name)] = value
                    if not df_grouped.empty:
                        all_data.append(df_grouped)
                    if log_area: log_area.write("✅ 处理完成\n")
                except Exception as e:
                    if log_area: log_area.write(f"❌ 错误: {e}\n")
    if all_data:
        df_all = pd.concat(all_data)
        df_final = df_all.groupby(df_all.index).sum()
        out_path = os.path.join(folder_path, output_file)
        df_final.to_excel(out_path)
        if log_area: log_area.write(f"\n✅ 汇总完成，保存至：{output_file}\n")
        return out_path, all_values
    else:
        if log_area: log_area.write("❌ 没有找到有效数据，请检查 Excel 格式或列名\n")
        return None, None

def update_file_b(file_a_path, file_b_path, log_area=None):
    try:
        df_a = pd.read_excel(file_a_path, index_col=0)
        wage_cols = df_a.columns
        all_values = {}
        for budget_unit, row in df_a.iterrows():
            for wage_type in wage_cols:
                value = row[wage_type]
                if "绩效工资" in wage_type:
                    wage_type = wage_type.replace("绩效工资", "基础性绩效")
                if "行政医疗" in wage_type:
                    wage_type = wage_type.replace("行政医疗", "职工基本医疗（行政）")
                elif "事业医疗" in wage_type:
                    wage_type = wage_type.replace("事业医疗", "基本医疗（事业）")
                elif "医疗保险" in wage_type:
                    wage_type = wage_type.replace("医疗保险", "基本医疗")
                key = (str(budget_unit).strip(), str(wage_type).strip())
                all_values[key] = value

        wb = load_workbook(file_b_path)
        sheet = wb.active
        j_col_index = 10
        match_count = 0
        for row_idx in range(2, sheet.max_row + 1):
            unit_cell = sheet.cell(row=row_idx, column=1)
            unit_info = str(unit_cell.value).strip() if unit_cell.value else ""
            budget_cell = sheet.cell(row=row_idx, column=2)
            budget_project = str(budget_cell.value).strip() if budget_cell.value else ""
            unit_info_cleaned = unit_info.replace("-", "").replace(" ", "")
            matched = False
            for (budget_unit, wage_type), value in all_values.items():
                budget_unit_cleaned = budget_unit.replace("-", "").replace(" ", "")
                unit_match = (budget_unit_cleaned in unit_info_cleaned) or (unit_info_cleaned in budget_unit_cleaned)
                wage_match = wage_type in budget_project
                if unit_match and wage_match:
                    sheet.cell(row=row_idx, column=j_col_index).value = value
                    match_count += 1
                    matched = True
                    if log_area: log_area.write(f"匹配成功: 行{row_idx} 单位:'{unit_info}' 项目:'{budget_project}' 值:{value}\n")
                    break
            if not matched and row_idx < 7:
                if log_area: log_area.write(f"未匹配: 行{row_idx} 单位:'{unit_info}' 项目:'{budget_project}'\n")
        output_path = os.path.join(os.path.dirname(file_b_path), "updated_" + os.path.basename(file_b_path))
        wb.save(output_path)
        if log_area: log_area.write(f"\n总共完成 {match_count} 处匹配\n")
        if log_area: log_area.write(f"已保存更新后的文件到: {output_path}\n")
        return output_path
    except Exception as e:
        if log_area: log_area.write(f"\n更新文件B出错: {e}\n")
        return None

# --------- Streamlit UI ---------

st.title("工资表自动处理工具（两步）")

uploaded_zip = st.file_uploader("第一步：上传工资表压缩包（.zip）", type=["zip"])
file_a_path = None
all_values = None

if uploaded_zip:
    with tempfile.TemporaryDirectory() as tmpdir:
        zip_path = os.path.join(tmpdir, "upload.zip")
        with open(zip_path, "wb") as f:
            f.write(uploaded_zip.read())

        with zipfile.ZipFile(zip_path, 'r') as z:
            z.extractall(tmpdir)

        st.markdown("### 📂 解压内容：")
        for root, _, files in os.walk(tmpdir):
            for f in files:
                st.markdown(f"- `{os.path.join(root, f).replace(tmpdir, '')}`")

        st.markdown("---")
        log_area = st.empty()
        st.markdown("### 🔧 处理中，生成文件A...")
        with st.spinner("处理中..."):
            file_a_path, all_values = process_file_a(tmpdir, log_area=log_area)

        if file_a_path:
            st.success("✅ 文件A已生成")

            st.markdown("---")
            uploaded_file_b = st.file_uploader("第二步：上传文件B（模板Excel）", type=["xlsx"])
            if uploaded_file_b:
                temp_b_path = os.path.join(tmpdir, "file_b.xlsx")
                with open(temp_b_path, "wb") as f:
                    f.write(uploaded_file_b.read())
                st.markdown("### 🔧 正在更新文件B...")
                with st.spinner("处理中..."):
                    updated_b_path = update_file_b(file_a_path, temp_b_path, log_area=log_area)
                if updated_b_path:
                    with open(updated_b_path, "rb") as f:
                        st.download_button(
                            label="📥 下载更新后的文件B",
                            data=f,
                            file_name="updated_文件B.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        else:
            st.error("❌ 文件A处理失败，请检查Excel格式或列名")
